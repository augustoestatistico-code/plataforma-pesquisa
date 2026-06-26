import os
import json
import pandas as pd
from sqlalchemy import create_engine, text
import dash
from dash import dcc, html, Input, Output, dash_table
import plotly.express as px
from flask import Flask, request, session, redirect, Response
import subprocess
import sys
import requests
from requests.auth import HTTPBasicAuth
import urllib.parse

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import io

# =========================
# CONFIG
# =========================

DATABASE_URL = os.getenv("DATABASE_URL", "").strip()

SECRET_KEY = os.getenv("SECRET_KEY", "123456")

server = Flask(__name__)
server.secret_key = SECRET_KEY

app = dash.Dash(
    __name__,
    server=server,
    url_base_pathname="/dashboard/",
    suppress_callback_exceptions=True
)

engine = create_engine(DATABASE_URL, pool_pre_ping=True)

# =========================
# LOGIN
# =========================
@server.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email")
        senha = request.form.get("senha")

        query = text("""
            SELECT id, cliente_id
            FROM usuarios
            WHERE email = :email AND senha = :senha
        """)

        with engine.connect() as conn:
            user = conn.execute(query, {"email": email, "senha": senha}).fetchone()

        if user:
            session["user_id"] = user[0]
            session["cliente_id"] = user[1]
            return redirect("/dashboard/")

        return """
        <body style="background:#0f172a;color:white;font-family:Arial;text-align:center;margin-top:100px">
            <h2>Login inválido</h2>
            <a href="/" style="color:#38bdf8">Voltar</a>
        </body>
        """

    return """
    <body style="margin:0;background:#0f172a;font-family:Arial;color:white">
        <div style="width:360px;margin:120px auto;background:#111827;padding:35px;border-radius:18px;box-shadow:0 0 25px #000">
            <h2 style="text-align:center">📊 Plataforma Pesquisa</h2>
            <form method="post">
                <label>Email</label>
                <input name="email" style="width:100%;padding:12px;margin:8px 0 18px;border-radius:8px;border:0">
                <label>Senha</label>
                <input name="senha" type="password" style="width:100%;padding:12px;margin:8px 0 22px;border-radius:8px;border:0">
                <button style="width:100%;padding:13px;background:#2563eb;color:white;border:0;border-radius:10px;font-weight:bold">
                    Entrar
                </button>
            </form>
        </div>
    </body>
    """


@server.route("/logout")
def logout():
    session.clear()
    return redirect("/")


@server.before_request
def proteger_dashboard():
    if request.path.startswith("/dashboard") and "cliente_id" not in session:
        return redirect("/")


# =========================
# FUNÇÕES
# =========================
def get_cliente(cliente_id):
    query = text("""
        SELECT nome, logo_url
        FROM clientes
        WHERE id = :cliente_id
    """)

    with engine.connect() as conn:
        row = conn.execute(query, {"cliente_id": cliente_id}).fetchone()

    if row:
        logo = row[1] or "/dashboard/assets/logos/sem_logo.png"
        return {"nome": row[0], "logo": logo}

    return {
        "nome": "Cliente",
        "logo": "/dashboard/assets/logos/sem_logo.png"
    }


def lista_pesquisas(cliente_id):
    df = pd.read_sql(
        text("""
            SELECT id, nome
            FROM pesquisas
            WHERE cliente_id = :cliente_id
            ORDER BY id DESC
        """),
        engine,
        params={"cliente_id": cliente_id}
    )

    return [{"label": row["nome"], "value": row["id"]} for _, row in df.iterrows()]


def carregar_dados(pesquisa_id):
    df = pd.read_sql(
        text("""
            SELECT id, submission_id, pesquisa_id, sexo, idade, localidade, dados, entrevistador
            FROM entrevistas
            WHERE pesquisa_id = :pesquisa_id
            ORDER BY id DESC
        """),
        engine,
        params={"pesquisa_id": pesquisa_id}
    )

    if df.empty:
        return df

    df["sexo"] = df["sexo"].fillna("Não informado").astype(str).str.strip().str.title()
    df["idade"] = df["idade"].fillna("Não informado").astype(str).str.strip()
    df["localidade"] = df["localidade"].fillna("Não informado").astype(str).str.strip().str.upper()
    df["entrevistador"] = df["entrevistador"].fillna("Não informado").astype(str).str.strip()

    return df


def normalizar_json(valor):
    if isinstance(valor, dict):
        return valor

    if isinstance(valor, str):
        try:
            return json.loads(valor)
        except Exception:
            return {}

    return {}

# =========================
# MOTOR DE INTELIGÊNCIA
# =========================

SEM_DIRECAO = {

    "",
    "NS",
    "NS/NR",
    "NÃO SABE",
    "NAO SABE",
    "NÃO SABE OPINAR",
    "NAO SABE OPINAR",
    "NÃO RESPONDEU",
    "NAO RESPONDEU",
    "NÃO QUIS RESPONDER",
    "NAO QUIS RESPONDER",
    "NÃO QUIS OPINAR",
    "NAO QUIS OPINAR",
    "SEM OPINIÃO",
    "SEM OPINIAO",
    "INDECISO",
    "INDECISA",
    "INDECISOS",
    "BRANCO",
    "NULO",
    "BRANCO/NULO",
    "NENHUM",
    "NENHUMA",
    "OUTRO",
    "OUTROS",
    "NR"
}

def eh_sem_direcao(valor):

    if pd.isna(valor):
        return True

    txt = str(valor).strip().upper()

    return txt in SEM_DIRECAO

def pegar_perfil_predominante(df_filtrado, coluna):

    if coluna not in df_filtrado.columns or df_filtrado.empty:
        return "Não informado", 0

    serie = df_filtrado[coluna].dropna().astype(str).str.strip()

    if serie.empty:
        return "Não informado", 0

    contagem = serie.value_counts()
    valor = contagem.idxmax()
    pct = round((contagem.max() / len(serie)) * 100, 1)

    return valor, pct


def perfil_resposta(df, pergunta, resposta):

    if "dados" not in df.columns:
        return None

    dados_json = df["dados"].apply(normalizar_json)
    json_df = pd.json_normalize(dados_json)

    coluna_real = None

    for c in json_df.columns:
        if c.upper() == pergunta.upper():
            coluna_real = c
            break

    if coluna_real is None:
        return None

    mascara = json_df[coluna_real].astype(str).str.strip().str.upper() == str(resposta).strip().upper()

    df_resp = df[mascara].copy()

    if df_resp.empty:
        return None

    sexo, sexo_pct = pegar_perfil_predominante(df_resp, "sexo")
    idade, idade_pct = pegar_perfil_predominante(df_resp, "idade")
    localidade, localidade_pct = pegar_perfil_predominante(df_resp, "localidade")

    escolaridade, escolaridade_pct = pegar_perfil_predominante(df_resp, "escolaridade")
    renda, renda_pct = pegar_perfil_predominante(df_resp, "renda_familiar")
    religiao, religiao_pct = pegar_perfil_predominante(df_resp, "religiao")
    zona, zona_pct = pegar_perfil_predominante(df_resp, "zona")

    return {
        "total": len(df_resp),
        "sexo": sexo,
        "sexo_pct": sexo_pct,
        "idade": idade,
        "idade_pct": idade_pct,
        "localidade": localidade,
        "localidade_pct": localidade_pct,
        "escolaridade": escolaridade,
        "escolaridade_pct": escolaridade_pct,
        "renda": renda,
        "renda_pct": renda_pct,
        "religiao": religiao,
        "religiao_pct": religiao_pct,
        "zona": zona,
        "zona_pct": zona_pct,
    }

def calcular_desvios(df_filtrado, df_total, coluna):

    if coluna not in df_total.columns:
        return []

    resultados = []

    total_base = len(df_total)
    total_filtrado = len(df_filtrado)

    if total_base == 0 or total_filtrado == 0:
        return []

    categorias = set(
        df_total[coluna].dropna().astype(str)
    )

    for categoria in categorias:

        pct_total = (
            (df_total[coluna].astype(str) == categoria).sum()
            / total_base
        ) * 100

        pct_filtrado = (
            (df_filtrado[coluna].astype(str) == categoria).sum()
            / total_filtrado
        ) * 100

        desvio = round(
            pct_filtrado - pct_total,
            1
        )

        resultados.append(
            (categoria, desvio)
        )

    resultados.sort(
        key=lambda x: x[1],
        reverse=True
    )

    return resultados

def grupos_fortes(df, pergunta, resposta):

    dados_json = df["dados"].apply(normalizar_json)
    json_df = pd.json_normalize(dados_json)

    coluna_real = None

    for c in json_df.columns:
        if c.upper() == pergunta.upper():
            coluna_real = c
            break

    if coluna_real is None:
        return []

    mascara = (
        json_df[coluna_real]
        .astype(str)
        .str.upper()
        .str.strip()
        ==
        str(resposta).upper().strip()
    )

    df_resp = df[mascara]

    destaques = []

    for coluna in [
        "sexo",
        "idade",
        "localidade",
        "zona",
        "escolaridade",
        "renda_familiar",
        "religiao"
    ]:
        desvios = calcular_desvios(
            df_resp,
            df,
            coluna
        )

        if len(desvios):

            categoria, valor = desvios[0]

            if valor >= 5:

                destaques.append(
                    f"{categoria} (+{valor} pts)"
                )

    return destaques

def grupos_fortes_fracos(df, pergunta, resposta):

    dados_json = df["dados"].apply(normalizar_json)
    json_df = pd.json_normalize(dados_json)

    coluna_real = None

    for c in json_df.columns:
        if c.upper() == pergunta.upper():
            coluna_real = c
            break

    if coluna_real is None:
        return [], []

    mascara = (
        json_df[coluna_real]
        .astype(str)
        .str.upper()
        .str.strip()
        ==
        str(resposta).upper().strip()
    )

    df_resp = df[mascara]

    fortes = []
    fracos = []

    for coluna in [
        "sexo",
        "idade",
        "localidade"
    ]:

        desvios = calcular_desvios(
            df_resp,
            df,
            coluna
        )

        for categoria, valor in desvios:

            if valor >= 5:
                fortes.append(
                    f"{categoria} (+{valor} pts)"
                )

            elif valor <= -5:
                fracos.append(
                    f"{categoria} ({valor} pts)"
                )

    return fortes[:3], fracos[:3]

def normalizar_resposta(valor):

    if pd.isna(valor):
        return ""

    txt = str(valor).strip()

    if txt == "":
        return ""

    return txt.title()


# =========================
# CLASSIFICAÇÃO DE PERGUNTAS
# =========================

def classificar_pergunta(label):

    txt = str(label).upper()

    PERFIL = [
        "SEXO",
        "IDADE",
        "ESCOLARIDADE",
        "RENDA",
        "RENDA_FAMILIAR",
        "RELIGIAO",
        "RELIGIÃO",
        "ZONA",
        "LOCALIDADE",
        "BAIRRO",
        "REGIAO",
        "REGIÃO"
    ]

    ESTRATEGICAS = [
        "PREFEITO",
        "GOVERNADOR",
        "PRESIDENTE",
        "DEPUTADO",
        "SENADOR",
        "VOTO",
        "ESPONT",
        "ESTIMUL",
        "REJEI",
        "EM QUEM VOTARIA",
        "INTENÇÃO",
        "INTENCAO"
    ]

    ANALITICAS = [
        "AVALIA",
        "APROVA",
        "DESAPROVA",
        "PROBLEMA",
        "PRIORIDADE",
        "SAUDE",
        "SAÚDE",
        "EDUCAC",
        "SEGURAN",
        "TRANSPORTE",
        "OBRAS",
        "LIMPEZA"
    ]

    if any(x in txt for x in PERFIL):
        return "PERFIL"

    if any(x in txt for x in ESTRATEGICAS):
        return "ESTRATEGICA"

    if any(x in txt for x in ANALITICAS):
        return "ANALITICA"

    return "GERAL"

def classificar_sentimento_resposta(valor):

    txt = str(valor).strip().upper()

    positivos = [
        "APROVA",
        "ÓTIMA",
        "OTIMA",
        "BOA",
        "BOM",
        "SIM",
        "SATISFEITO",
        "POSITIVO"
    ]

    negativos = [
        "DESAPROVA",
        "REPROVA",
        "RUIM",
        "PÉSSIMA",
        "PESSIMA",
        "NÃO",
        "NAO",
        "INSATISFEITO",
        "NEGATIVO"
    ]

    neutros = [
        "REGULAR",
        "MAIS OU MENOS",
        "NEUTRO"
    ]

    if any(x in txt for x in positivos):
        return "POSITIVO"

    if any(x in txt for x in negativos):
        return "NEGATIVO"

    if any(x in txt for x in neutros):
        return "NEUTRO"

    return "OUTRO"


import unicodedata
import re

def normalizar_texto_aberto(texto):

    if pd.isna(texto):
        return ""

    texto = str(texto).lower()

    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ascii", "ignore").decode("utf-8")

    texto = re.sub(r"[^a-z0-9 ]", " ", texto)
    texto = re.sub(r"\s+", " ", texto)

    return texto.strip()


def extrair_gps(df):
    pontos = []

    if "dados" not in df.columns:
        return pd.DataFrame()

    for _, row in df.iterrows():
        dados = normalizar_json(row.get("dados"))

        lat = dados.get("latitude_melhor") or dados.get("lat_final") or dados.get("lat_inicio")
        lon = dados.get("longitude_melhor") or dados.get("lon_final") or dados.get("lon_inicio")

        if (not lat or not lon) and isinstance(dados.get("gps_final"), dict):
            coords = dados["gps_final"].get("coordinates", [])
            if len(coords) >= 2:
                lon = coords[0]
                lat = coords[1]

        try:
            if lat and lon:
                pontos.append({
                    "submission_id": row.get("submission_id"),
                    "lat": float(lat),
                    "lon": float(lon),
                    "localidade": row.get("localidade", ""),
                    "entrevistador": row.get("entrevistador", ""),
                    "accuracy": dados.get("accuracy_melhor") or dados.get("acc_final") or ""
                })
        except:
            pass

    return pd.DataFrame(pontos)


def card(titulo, valor, subtitulo="", cor="#2563eb"):
    return html.Div([
        html.Div(titulo.upper(), style={
            "fontSize": "12px",
            "fontWeight": "bold",
            "color": "#bfdbfe"
        }),
        html.Div(valor, style={
            "fontSize": "34px",
            "fontWeight": "bold",
            "marginTop": "8px",
            "color": "white"
        }),
        html.Div(subtitulo, style={
            "fontSize": "12px",
            "color": "#cbd5e1",
            "marginTop": "6px"
        })
    ], style={
        "background": f"linear-gradient(135deg, {cor}, #0f172a)",
        "padding": "22px",
        "borderRadius": "16px",
        "border": "1px solid #1e293b",
        "boxShadow": "0 8px 24px rgba(0,0,0,.35)"
    })

def tema_fig(fig):
    fig.update_layout(
        paper_bgcolor="#111827",
        plot_bgcolor="#111827",
        font_color="#e5e7eb",
        title_font_color="#ffffff",
        margin=dict(l=35, r=25, t=55, b=35),
        legend=dict(bgcolor="rgba(0,0,0,0)")
    )

    fig.update_xaxes(gridcolor="#1f2937")
    fig.update_yaxes(gridcolor="#1f2937")

    return fig

def cor_resposta_mapa(valor):

    valor = str(valor).strip().lower()

    mapa = {
        "aprova": "#16a34a",
        "desaprova": "#dc2626",
        "ótima": "#15803d",
        "otima": "#15803d",
        "boa": "#22c55e",
        "regular": "#facc15",
        "ruim": "#f97316",
        "péssima": "#b91c1c",
        "pessima": "#b91c1c",
        "sim": "#16a34a",
        "não": "#dc2626",
        "nao": "#dc2626",
        "talvez": "#facc15",
        "indeciso": "#9ca3af",
        "indeciso ": "#9ca3af",
        "não sabe": "#9ca3af",
        "nao sabe": "#9ca3af",
        "ns/nr": "#9ca3af",
        "nenhum": "#6b7280",
        "branco/nulo": "#6b7280",
    }

    return mapa.get(valor, "#3b82f6")

    fig.update_xaxes(gridcolor="#1f2937")
    fig.update_yaxes(gridcolor="#1f2937")
    return fig


def pergunta_deve_ignorar(coluna):
    colunas_ignorar_fixas = {
        "meta.instanceID",
        "instanceID",
        "__system",
        "_id",
        "_uuid",
        "_submission_time",
        "nome",
        "loc1",
        "pesquisa_inicio",
        "pesquisa_fim",
        "hoje",
        "HOJE",
        "PESQUISA_INICIO",
        "PESQUISA_FIM"
    }

    if coluna in colunas_ignorar_fixas:
        return True

    if coluna.startswith("_system."):
        return True

    if coluna.startswith("_"):
        return True

    return False

def gerar_graficos_perguntas(df, pesquisa_id, pergunta_selecionada=None):
    perguntas = []

    if "dados" not in df.columns:
        return perguntas

    dados_normalizados = df["dados"].apply(normalizar_json)
    json_df = pd.json_normalize(dados_normalizados)

    if json_df.empty:
        return perguntas

    perguntas_exibir = pd.read_sql(
        text("""
            SELECT
                UPPER(name) AS name,
                label,
                id
            FROM perguntas_pesquisa
            WHERE pesquisa_id = :pesquisa_id
            AND exibir_dashboard = true
            ORDER BY id
        """),
        engine,
        params={"pesquisa_id": pesquisa_id}
    )

    lista_exibir = set(perguntas_exibir["name"].tolist())
    labels = dict(zip(perguntas_exibir["name"], perguntas_exibir["label"]))

    if pergunta_selecionada:
        ordem_perguntas = [pergunta_selecionada.upper()]
    else:
        ordem_perguntas = perguntas_exibir["name"].tolist()

    for coluna_upper in ordem_perguntas:

        coluna_real = None

        for c in json_df.columns:
            if c.upper() == coluna_upper:
                coluna_real = c
                break

        if coluna_real is None:
            continue

        coluna = coluna_real
        coluna_upper = coluna.upper()

        if coluna_upper not in lista_exibir:
            continue

        serie = json_df[coluna].dropna().astype(str).str.strip()
        serie = serie[serie != ""]

        if serie.empty:
            continue

        contagem = serie.value_counts().reset_index()
        contagem.columns = ["Resposta", "Quantidade"]
        contagem["%"] = (contagem["Quantidade"] / contagem["Quantidade"].sum() * 100).round(1)
        contagem["Texto"] = (
            contagem["Quantidade"].astype(str)
            + " ("
            + contagem["%"].astype(str)
            + "%)"
        )

        titulo = labels.get(coluna_upper, coluna)

        fig = px.bar(
            contagem.sort_values("Quantidade", ascending=True),
            x="Quantidade",
            y="Resposta",
            orientation="h",
            text="Texto",
            title=titulo
        )

        fig = tema_fig(fig)

        tabela_respostas = dash_table.DataTable(
            data=contagem.to_dict("records"),
            columns=[
                {"name": "Resposta", "id": "Resposta"},
                {"name": "Quantidade", "id": "Quantidade"},
                {"name": "%", "id": "%"},
            ],
            style_table={"overflowX": "auto"},
            style_header={
                "backgroundColor": "#020617",
                "color": "white",
                "fontWeight": "bold",
                "border": "1px solid #1f2937"
            },
            style_cell={
                "backgroundColor": "#111827",
                "color": "#e5e7eb",
                "border": "1px solid #1f2937",
                "padding": "8px",
                "textAlign": "left"
            },
            page_size=15
        )

        perguntas.append(
            html.Div([
                dcc.Graph(figure=fig),
                html.H4("Tabela de respostas", style={"marginTop": "10px"}),
                tabela_respostas
            ], style={
                "background": "#111827",
                "borderRadius": "18px",
                "border": "1px solid #1f2937",
                "marginBottom": "18px",
                "padding": "14px"
            })
        )

    return perguntas

# =========================
# ÁUDIOS
# =========================
def carregar_audios(pesquisa_id):

    try:
        sql = """
            SELECT
                pesquisa_id,
                submission_id,
                entrevistador,
                localidade,
                data_entrevista,
                nome_arquivo
            FROM audios_entrevistas
            WHERE pesquisa_id = %(pesquisa_id)s
            ORDER BY id DESC
            LIMIT 100
        """

        df = pd.read_sql(sql, engine, params={"pesquisa_id": pesquisa_id})

        if df.empty:
            return html.Div("Nenhum áudio disponível nesta pesquisa.", style={"color": "#9ca3af"})

        linhas = []

        for _, row in df.iterrows():
            audio_src = f"/audio/{row['pesquisa_id']}/{row['submission_id']}/{row['nome_arquivo']}"

            linhas.append(html.Tr([
                html.Td(row["entrevistador"]),
                html.Td(row["localidade"]),
                html.Td(str(row["data_entrevista"])),
                html.Td( html.Audio(
                            src=audio_src,
                            controls=True,
                            style={"width": "300px"}
                        ))
            ]))

        return html.Table([
            html.Thead(html.Tr([
                html.Th("Entrevistador"),
                html.Th("Localidade"),
                html.Th("Data"),
                html.Th("Áudio")
            ])),
            html.Tbody(linhas)
        ], style={"width": "100%", "color": "white"})

    except Exception as e:
        return html.Div(f"Erro ao carregar áudios: {e}")
    
# =========================
# LAYOUT
# =========================

# =========================
# LAYOUT POWER BI
# =========================
app.layout = html.Div([
    dcc.Location(id="url"),

    dcc.Interval(
        id="interval-atualizacao",
        interval=15 * 60 * 1000,  # 15 minutos
        n_intervals=0
    ),    

    # SIDEBAR
    html.Div([
        html.H2([
            html.Span("Ipsensus", style={"color": "#1e88ff"}),
            html.Span(" Survey", style={"color": "white"})
        ], style={"margin": "0", "fontSize": "24px"}),

        html.Div("MENU", style={
            "color": "#94a3b8",
            "fontSize": "13px",
            "marginTop": "30px",
            "marginBottom": "12px"
        }),

        html.Button("📊 Visão Geral", id="btn-visao-geral", n_clicks=0, style={
            "display": "block", "width": "100%", "textAlign": "left",
            "background": "#2563eb", "padding": "12px",
            "borderRadius": "8px", "marginBottom": "10px",
            "fontWeight": "bold", "color": "white",
            "border": "0", "cursor": "pointer"
        }),

        html.Button("📝 Entrevistas", id="btn-entrevistas", n_clicks=0, style={
            "display": "block", "width": "100%", "textAlign": "left",
            "padding": "10px", "color": "#cbd5e1",
            "background": "transparent", "border": "0", "cursor": "pointer"
        }),

        html.Button("🗺️ Mapas", id="btn-mapas", n_clicks=0, style={
            "display": "block", "width": "100%", "textAlign": "left",
            "padding": "10px", "color": "#cbd5e1",
            "background": "transparent", "border": "0", "cursor": "pointer"
        }),

        html.Button("📋 Perguntas", id="btn-perguntas", n_clicks=0, style={
            "display": "block", "width": "100%", "textAlign": "left",
            "padding": "10px", "color": "#cbd5e1",
            "background": "transparent", "border": "0", "cursor": "pointer"
        }),

        html.Button("👥 Entrevistadores", id="btn-entrevistadores", n_clicks=0, style={
            "display": "block", "width": "100%", "textAlign": "left",
            "padding": "10px", "color": "#cbd5e1",
            "background": "transparent", "border": "0", "cursor": "pointer"
        }),

        html.Button("📄 Relatórios", id="btn-relatorios", n_clicks=0, style={
            "display": "block", "width": "100%", "textAlign": "left",
            "padding": "10px", "color": "#cbd5e1",
            "background": "transparent", "border": "0", "cursor": "pointer"
        }),
        html.Button("🧠 Inteligência", id="btn-inteligencia", n_clicks=0, style={
            "display": "block", "width": "100%", "textAlign": "left",
            "padding": "10px", "color": "#cbd5e1",
            "background": "transparent", "border": "0", "cursor": "pointer"
        }),
        
        html.Hr(style={"borderColor": "#1f2937", "marginTop": "25px"}),

        html.Div("FILTROS", style={
            "color": "#94a3b8",
            "fontSize": "13px",
            "marginBottom": "12px"
        }),

        html.Label("Pesquisa", style={"fontSize": "13px"}),
        dcc.Dropdown(
            id="pesquisa",
            options=[],
            value=None,
            placeholder="Selecione",
            style={"color": "#111827", "marginTop": "6px", "marginBottom": "16px"}
        ),

        html.Label("Localidade", style={"fontSize": "13px"}),
        dcc.Dropdown(
            id="filtro-localidade",
            options=[],
            value=None,
            placeholder="Todas",
            style={"color": "#111827", "marginTop": "6px", "marginBottom": "16px"}
        ),

        html.Label("Entrevistador", style={"fontSize": "13px"}),
        dcc.Dropdown(
            id="filtro-entrevistador",
            options=[],
            value=None,
            placeholder="Todos",
            style={"color": "#111827", "marginTop": "6px", "marginBottom": "16px"}
        ),

        html.Label("Pergunta para análise", style={"fontSize": "13px"}),
        dcc.Dropdown(
            id="pergunta-mapa",
            options=[],
            value=None,
            placeholder="Selecione uma pergunta",
            style={"color": "#111827", "marginTop": "6px", "marginBottom": "16px"}
        ),

        html.Label("Tipo de Mapa", style={"fontSize": "13px"}),
        dcc.Dropdown(
            id="tipo-mapa",
            options=[
                {"label": "Mapa de Pontos", "value": "pontos"},
                {"label": "Mapa de Calor", "value": "calor"},
            ],
            value="pontos",
            clearable=False,
            style={"color": "#111827", "marginTop": "6px", "marginBottom": "16px"}
        ),

        html.A("Sair", href="/logout", style={
            "display": "block",
            "marginTop": "25px",
            "color": "white",
            "background": "#dc2626",
            "padding": "10px",
            "borderRadius": "8px",
            "textAlign": "center",
            "textDecoration": "none"
        }),

    ], style={
        "position": "fixed",
        "left": "0",
        "top": "0",
        "bottom": "0",
        "width": "260px",
        "background": "#020617",
        "padding": "24px 16px",
        "borderRight": "1px solid #1f2937",
        "color": "white",
        "overflowY": "auto"
    }),

    # CONTEÚDO PRINCIPAL
        html.Div([

            html.Div([
                html.Div([
                    html.H2("Acompanhamento em Tempo Real", style={"margin": "0"}),
                    html.Div("Dados atualizados automaticamente", style={
                        "color": "#94a3b8",
                        "fontSize": "13px",
                        "marginTop": "4px"
                    }),
                ]),

                html.Div([
                    html.Div(id="cliente-header", style={"color": "#cbd5e1"}),
                    html.Div(id="pesquisa-header", style={
                        "color": "#94a3b8",
                        "fontSize": "12px",
                        "marginTop": "6px",
                        "textAlign": "right"
                    })
                ])
            ], style={
                "display": "flex",
                "justifyContent": "space-between",
                "alignItems": "center",
                "marginBottom": "22px"
            }),

        html.Div([
            html.Div(id="kpis", style={
                "display": "grid",
                "gridTemplateColumns": "repeat(4, 1fr)",
                "gap": "18px",
                "marginBottom": "20px"
            }),

            html.Div([
                html.Div([dcc.Graph(id="grafico-sexo")], style={
                    "background": "#111827",
                    "borderRadius": "14px",
                    "border": "1px solid #1f2937",
                    "padding": "10px"
                }),
                html.Div([dcc.Graph(id="grafico-idade")], style={
                    "background": "#111827",
                    "borderRadius": "14px",
                    "border": "1px solid #1f2937",
                    "padding": "10px"
                }),
            ], style={
                "display": "grid",
                "gridTemplateColumns": "1fr 1fr",
                "gap": "18px",
                "marginBottom": "20px"
            }),
        ], id="secao-visao-geral"),

        html.Div([
            html.Div([dcc.Graph(id="mapa-gps")], style={
                "background": "#111827",
                "borderRadius": "14px",
                "border": "1px solid #1f2937",
                "padding": "10px"
            }),
        ], id="secao-mapas"),

        html.Div([
            html.H2("Resultados das Perguntas", style={"marginBottom": "16px"}),
            html.Div(id="perguntas-dinamicas")
        ], id="secao-perguntas"),

        html.Div([
            html.H3("Desempenho por Entrevistador", style={"marginTop": "0"}),
            html.Div(id="tabela-entrevistador")
        ], id="secao-entrevistadores", style={
            "background": "#111827",
            "borderRadius": "14px",
            "border": "1px solid #1f2937",
            "padding": "18px"
        }),

        html.Div([
            html.H2("Entrevistas / Auditoria", style={
                "marginTop": "0",
                "marginBottom": "16px"
            }),
            html.H3("Lista de Entrevistas"),
            html.Div(id="tabela-entrevistas", style={"marginBottom": "25px"}),

            html.H3("Auditoria de Áudios"),
            html.Div(id="audios-entrevistas")
        ], id="secao-entrevistas", style={
            "background": "#111827",
            "borderRadius": "14px",
            "border": "1px solid #1f2937",
            "padding": "20px",
            "marginTop": "20px"
        }),

        html.Div([
            html.H2("Relatórios", style={
                "marginTop": "0",
                "marginBottom": "16px"
            }),

            html.Div([
                html.A(
                    "📥 Exportar base da pesquisa em Excel",
                    id="link-exportar-excel",
                    href="#",
                    target="_blank",
                    style={
                        "display": "inline-block",
                        "background": "#2563eb",
                        "color": "white",
                        "padding": "12px 18px",
                        "borderRadius": "10px",
                        "textDecoration": "none",
                        "fontWeight": "bold"
                    }
                ),

            

                html.A(
                    "📄 Gerar Relatório PDF",
                    id="link-gerar-pdf",
                    href="#",
                    target="_blank",
                    style={
                        "display": "inline-block",
                        "background": "#16a34a",
                        "color": "white",
                        "padding": "12px 18px",
                        "borderRadius": "10px",
                        "textDecoration": "none",
                        "fontWeight": "bold",
                        "marginLeft": "10px"
                    }
                ),

                html.A(
                    "📊 Exportar tabelas % e amostra",
                    id="link-exportar-tabelas",
                    href="#",
                    target="_blank",
                    style={
                        "display": "inline-block",
                        "background": "#7c3aed",
                        "color": "white",
                        "padding": "12px 18px",
                        "borderRadius": "10px",
                        "textDecoration": "none",
                        "fontWeight": "bold",
                        "marginLeft": "10px",
                        "marginTop": "10px"
                    }
                )

            ], style={
                "background": "#111827",
                "borderRadius": "14px",
                "border": "1px solid #1f2937",
                "padding": "20px",
                "color": "#cbd5e1"
            })
        ], id="secao-relatorios"),
        html.Div([

            html.H2(
                "Inteligência Eleitoral",
                style={"marginBottom": "20px"}
            ),

            html.Div(
                id="inteligencia-conteudo"
            )

        ],
        id="secao-inteligencia",
        style={
            "display": "none"
        }),
                    

    ], style={
        "marginLeft": "260px",
        "padding": "24px",
        "background": "#0f172a",
        "minHeight": "100vh",
        "color": "#e5e7eb",
        "fontFamily": "Arial, sans-serif"
    })

])

# =========================
# CALLBACK PESQUISAS
# =========================
@app.callback(
    [
        Output("pesquisa", "options"),
        Output("pesquisa", "value"),
        Output("cliente-header", "children"),
    ],
    Input("url", "pathname")
)
def inicializar_dashboard(pathname):
    cliente_id = session.get("cliente_id")

    if not cliente_id:
        return [], None, ""

    cliente = get_cliente(cliente_id)
    options = lista_pesquisas(cliente_id)
    valor_inicial = options[0]["value"] if options else None

    logo = cliente.get("logo")
    header = html.Div([
        html.Img(
            src=logo,
            style={
                "height": "45px",
                "marginRight": "12px",
                "borderRadius": "8px",
                "objectFit": "contain",
                "background": "white"
            }
        ) if logo else None,

        html.Div([
            html.Div(f"Cliente: {cliente['nome']}", style={
                "fontWeight": "bold",
                "fontSize": "15px"
            }),
            html.Div("Ipsensus Survey", style={
                "fontSize": "13px",
                "color": "#38bdf8",
                "fontWeight": "bold",
                "marginTop": "2px"
            }),
            html.Div("Dashboard de pesquisas em andamento", style={
                "fontSize": "12px",
                "color": "#94a3b8",
                "marginTop": "2px"
            })
        ])
    ], style={
        "display": "flex",
        "alignItems": "center",
        "gap": "10px"
    })

    return options, valor_inicial, header


# =========================
# CALLBACK DASHBOARD

@app.callback(
    [
        Output("kpis", "children"),
        Output("grafico-sexo", "figure"),
        Output("grafico-idade", "figure"),
        Output("mapa-gps", "figure"),
        Output("tabela-entrevistador", "children"),
        Output("tabela-entrevistas", "children"),
        Output("perguntas-dinamicas", "children"),
        Output("audios-entrevistas", "children"),
        Output("filtro-localidade", "options"),
        Output("filtro-entrevistador", "options"),
        Output("pergunta-mapa", "options"),
        Output("pesquisa-header", "children"),
    ],
    [
        Input("pesquisa", "value"),
        Input("filtro-localidade", "value"),
        Input("filtro-entrevistador", "value"),
        Input("pergunta-mapa", "value"),
        Input("tipo-mapa", "value"),
        Input("interval-atualizacao", "n_intervals"),
    ]
)
def atualizar_dashboard(pesquisa_id, filtro_localidade, filtro_entrevistador, pergunta_mapa, tipo_mapa, n_intervals):   
    if not pesquisa_id:
        fig_vazio = tema_fig(px.bar(title="Sem pesquisa selecionada"))
        return [], fig_vazio, fig_vazio, fig_vazio, "", "", "", "", [], [], [], ""

    df = carregar_dados(pesquisa_id)

    if df.empty:
        fig_vazio = tema_fig(px.bar(title="Sem dados"))
        return [
            card("Total", "0", "Sem entrevistas")
        ], fig_vazio, fig_vazio, fig_vazio, "Sem dados", "", "", "", [], [], [], ""

    opcoes_localidade = [
        {"label": x, "value": x}
        for x in sorted(df["localidade"].dropna().unique())
    ]

    opcoes_entrevistador = [
        {"label": x, "value": x}
        for x in sorted(df["entrevistador"].dropna().unique())
    ]

    if filtro_localidade:
        df = df[df["localidade"] == filtro_localidade]

    if filtro_entrevistador:
        df = df[df["entrevistador"] == filtro_entrevistador]

    if df.empty:
        fig_vazio = tema_fig(px.bar(title="Sem dados para o filtro selecionado"))
        return [
            card("Total", "0", "Filtro sem entrevistas")
        ], fig_vazio, fig_vazio, fig_vazio, "Sem dados", "", "", "", opcoes_localidade, opcoes_entrevistador, [], ""

    total = len(df)
    entrevistadores = df["entrevistador"].nunique()

    masc = int((df["sexo"] == "Masculino").sum())
    fem = int((df["sexo"] == "Feminino").sum())

    masc_pct = round((masc / total) * 100, 1) if total else 0
    fem_pct = round((fem / total) * 100, 1) if total else 0

    kpis = [
        card("Entrevistas realizadas", f"{total:,}".replace(",", "."), "Amostra filtrada", "#1d4ed8"),
        card("Masculino", f"{masc_pct}%", f"{masc} entrevistas", "#0f766e"),
        card("Feminino", f"{fem_pct}%", f"{fem} entrevistas", "#be185d"),
        card("Entrevistadores", entrevistadores, "Equipe em campo", "#7e22ce"),
    ]

    sexo_df = df["sexo"].value_counts().reset_index()
    sexo_df.columns = ["Sexo", "Quantidade"]

    fig_sexo = px.pie(
        sexo_df,
        names="Sexo",
        values="Quantidade",
        title="Distribuição por Sexo",
        hole=0.45
    )
    fig_sexo = tema_fig(fig_sexo)

    idade_ordem = [
        "16 a 24 anos",
        "25 a 34 anos",
        "35 a 44 anos",
        "45 a 59 anos",
        "60 anos ou mais",
        "Não informado"
    ]

    idade_df = df["idade"].value_counts().reset_index()
    idade_df.columns = ["Faixa Etária", "Quantidade"]
    idade_df["Faixa Etária"] = pd.Categorical(
        idade_df["Faixa Etária"],
        categories=idade_ordem,
        ordered=True
    )
    idade_df = idade_df.sort_values("Faixa Etária")

    fig_idade = px.bar(
        idade_df,
        x="Faixa Etária",
        y="Quantidade",
        text="Quantidade",
        title="Distribuição por Faixa Etária"
    )
    fig_idade = tema_fig(fig_idade)

    ent_df = df["entrevistador"].value_counts().reset_index()
    ent_df.columns = ["Entrevistador", "Quantidade"]
    ent_df["%"] = (ent_df["Quantidade"] / total * 100).round(1)

    tabela_ent = dash_table.DataTable(
        data=ent_df.to_dict("records"),
        columns=[{"name": c, "id": c} for c in ent_df.columns],
        style_table={"overflowX": "auto"},
        style_header={
            "backgroundColor": "#020617",
            "color": "white",
            "fontWeight": "bold",
            "border": "1px solid #1f2937"
        },
        style_cell={
            "backgroundColor": "#111827",
            "color": "#e5e7eb",
            "border": "1px solid #1f2937",
            "padding": "10px",
            "textAlign": "left"
        },
        page_size=10
    )

    tabela_entrevistas = dash_table.DataTable(
        data=df[[
            "id",
            "submission_id",
            "entrevistador",
            "localidade",
            "sexo",
            "idade"
        ]].head(300).to_dict("records"),
        columns=[
            {"name": "ID", "id": "id"},
            {"name": "Submission", "id": "submission_id"},
            {"name": "Entrevistador", "id": "entrevistador"},
            {"name": "Localidade", "id": "localidade"},
            {"name": "Sexo", "id": "sexo"},
            {"name": "Idade", "id": "idade"},
        ],
        page_size=15,
        filter_action="native",
        sort_action="native",
        style_table={"overflowX": "auto"},
        style_header={
            "backgroundColor": "#020617",
            "color": "white",
            "fontWeight": "bold",
            "border": "1px solid #1f2937"
        },
        style_cell={
            "backgroundColor": "#111827",
            "color": "#e5e7eb",
            "border": "1px solid #1f2937",
            "padding": "8px",
            "fontSize": "12px",
            "textAlign": "left",
            "maxWidth": "220px",
            "overflow": "hidden",
            "textOverflow": "ellipsis"
        },
    )    

    gps_df = extrair_gps(df)

    if not gps_df.empty and pergunta_mapa:
        mapa_respostas = {}

        for _, row in df.iterrows():
            dados = normalizar_json(row.get("dados"))
            submission = row.get("submission_id")
            mapa_respostas[submission] = dados.get(pergunta_mapa)

        gps_df["resposta_mapa"] = gps_df["submission_id"].map(mapa_respostas)

    if gps_df.empty:

        fig_mapa = tema_fig(
            px.scatter(title="Sem GPS disponível nesta pesquisa")
        )

    else:

        if tipo_mapa == "calor":

            fig_mapa = px.density_mapbox(
                gps_df,
                lat="lat",
                lon="lon",
                radius=35,
                zoom=12,
                height=650,
                title="Mapa de Calor das Entrevistas"
            )

        elif pergunta_mapa and "resposta_mapa" in gps_df.columns:

            gps_df["resposta_mapa"] = gps_df["resposta_mapa"].fillna("Não informado")

            fig_mapa = px.scatter_mapbox(
                gps_df,
                lat="lat",
                lon="lon",
                color="resposta_mapa",

                color_discrete_map={
                    resposta: cor_resposta_mapa(resposta)
                    for resposta in gps_df["resposta_mapa"].unique()
                },

                hover_name="localidade",
                hover_data=[
                    "entrevistador",
                    "accuracy",
                    "resposta_mapa"
                ],
                zoom=12,
                height=650,
                title="Mapa por Resposta da Pergunta"
            )
        else:

            fig_mapa = px.scatter_mapbox(
                gps_df,
                lat="lat",
                lon="lon",
                hover_name="localidade",
                hover_data=[
                    "entrevistador",
                    "accuracy"
                ],
                zoom=12,
                height=650,
                size=[14] * len(gps_df),
                title="Mapa de Pontos das Entrevistas"
            )

        fig_mapa.update_layout(
            mapbox_style="open-street-map",
            paper_bgcolor="#111827",
            plot_bgcolor="#111827",
            font_color="#e5e7eb",
            margin=dict(l=0, r=0, t=50, b=0)
        )

    perguntas = gerar_graficos_perguntas(df, pesquisa_id, pergunta_mapa)
    audios = carregar_audios(pesquisa_id)

    if not perguntas:
        perguntas = [
            html.Div("Nenhuma pergunta encontrada no campo dados JSONB.", style={
                "background": "#111827",
                "padding": "20px",
                "borderRadius": "18px"
            })
        ]
    perguntas_mapa_df = pd.read_sql(
        text("""
            SELECT UPPER(name) AS name, label
            FROM perguntas_pesquisa
            WHERE pesquisa_id = :pesquisa_id
            AND exibir_dashboard = true
            ORDER BY id
        """),
        engine,
        params={"pesquisa_id": pesquisa_id}
    )

    opcoes_pergunta_mapa = [
        {"label": row["label"], "value": row["name"]}
        for _, row in perguntas_mapa_df.iterrows()
    ]
    
    pesquisa_nome = ""

    try:
        pesquisa_nome = [
            opt["label"]
            for opt in lista_pesquisas(session.get("cliente_id"))
            if opt["value"] == pesquisa_id
        ][0]
    except:
        pesquisa_nome = f"Pesquisa ID {pesquisa_id}"

    pesquisa_header = html.Div([
        html.Div(f"Pesquisa: {pesquisa_nome}"),
        html.Div(f"Entrevistas filtradas: {total}")
    ])
    return (
        kpis,
        fig_sexo,
        fig_idade,
        fig_mapa,
        tabela_ent,
        tabela_entrevistas,
        perguntas,
        audios,
        opcoes_localidade,
        opcoes_entrevistador,
        opcoes_pergunta_mapa,
        pesquisa_header
    )

# =========================
# CALLBACK INTELIGÊNCIA
# =========================
@app.callback(
    Output("inteligencia-conteudo", "children"),
    [
        Input("pesquisa", "value"),
        Input("interval-atualizacao", "n_intervals"),
    ]
)
def gerar_inteligencia(pesquisa_id, n_intervals):

    try:
        if not pesquisa_id:
            return html.Div("Selecione uma pesquisa para gerar a inteligência.", style={"color": "#cbd5e1"})

        df = carregar_dados(pesquisa_id)

        if df.empty:
            return html.Div("Sem dados para gerar inteligência.", style={"color": "#cbd5e1"})

        total = len(df)

        sexo_top = df["sexo"].value_counts().idxmax() if "sexo" in df.columns and not df["sexo"].empty else "Não informado"
        idade_top = df["idade"].value_counts().idxmax() if "idade" in df.columns and not df["idade"].empty else "Não informado"

        pct_fem = round((df["sexo"].eq("Feminino").sum() / total) * 100, 1) if total and "sexo" in df.columns else 0
        pct_masc = round((df["sexo"].eq("Masculino").sum() / total) * 100, 1) if total and "sexo" in df.columns else 0

        sexo_dist = df["sexo"].value_counts().reset_index() if "sexo" in df.columns else pd.DataFrame(columns=["Sexo", "Entrevistas"])
        if not sexo_dist.empty:
            sexo_dist.columns = ["Sexo", "Entrevistas"]
            sexo_dist["%"] = (sexo_dist["Entrevistas"] / total * 100).round(1)

        idade_dist = df["idade"].value_counts().reset_index() if "idade" in df.columns else pd.DataFrame(columns=["Faixa Etária", "Entrevistas"])
        if not idade_dist.empty:
            idade_dist.columns = ["Faixa Etária", "Entrevistas"]
            idade_dist["%"] = (idade_dist["Entrevistas"] / total * 100).round(1)

        localidades_top = df["localidade"].value_counts().head(5).reset_index() if "localidade" in df.columns else pd.DataFrame(columns=["Localidade", "Entrevistas"])
        if not localidades_top.empty:
            localidades_top.columns = ["Localidade", "Entrevistas"]
            localidades_top["%"] = (localidades_top["Entrevistas"] / total * 100).round(1)

        top_localidades_txt = ", ".join([
            f"{row['Localidade']} ({row['%']}%)"
            for _, row in localidades_top.head(4).iterrows()
        ]) if not localidades_top.empty else "não informado"

        sexo_txt = ", ".join([
            f"{row['Sexo']} {row['%']}%"
            for _, row in sexo_dist.iterrows()
        ]) if not sexo_dist.empty else "não informado"

        idade_txt = ", ".join([
            f"{row['Faixa Etária']} {row['%']}%"
            for _, row in idade_dist.iterrows()
        ]) if not idade_dist.empty else "não informado"

        gps_df = extrair_gps(df)
        entrevistas_gps = len(gps_df)
        gps_pct = round((entrevistas_gps / total) * 100, 1) if total else 0

        if entrevistas_gps > 0:
            mapa_txt = (
                f"A base possui {entrevistas_gps} entrevistas com coordenadas GPS válidas "
                f"({gps_pct}% da amostra), permitindo análise territorial por pontos, "
                f"mapa de calor e cruzamento com respostas da pesquisa."
            )
        else:
            mapa_txt = (
                "A base não possui coordenadas GPS válidas suficientes para análise territorial por mapa."
            )

        def tabela_padrao(df_tabela, page_size=10):
            return dash_table.DataTable(
                data=df_tabela.to_dict("records"),
                columns=[{"name": c, "id": c} for c in df_tabela.columns],
                style_table={"overflowX": "auto"},
                style_header={
                    "backgroundColor": "#020617",
                    "color": "white",
                    "fontWeight": "bold",
                    "border": "1px solid #1f2937"
                },
                style_cell={
                    "backgroundColor": "#111827",
                    "color": "#e5e7eb",
                    "border": "1px solid #1f2937",
                    "padding": "9px",
                    "textAlign": "left"
                },
                page_size=page_size
            )

        tabela_localidades = tabela_padrao(localidades_top, 5)
        tabela_sexo = tabela_padrao(sexo_dist, 10)
        tabela_idade = tabela_padrao(idade_dist, 10)

        # =========================
        # INTELIGÊNCIA DAS PERGUNTAS
        # =========================
        if "dados" not in df.columns:
            json_df = pd.DataFrame()
        else:
            dados_json = df["dados"].apply(normalizar_json)
            json_df = pd.json_normalize(dados_json)

        perguntas_df = pd.read_sql(
            text("""
                SELECT UPPER(name) AS name, label
                FROM perguntas_pesquisa
                WHERE pesquisa_id = :pesquisa_id
                AND exibir_dashboard = true
                ORDER BY id
            """),
            engine,
            params={"pesquisa_id": pesquisa_id}
        )

        blocos_perguntas = []

        for _, pergunta in perguntas_df.iterrows():
            try:
                nome_coluna = str(pergunta["name"]).upper()
                label = pergunta["label"]
                tipo_pergunta = classificar_pergunta(label)

                if tipo_pergunta == "PERFIL":
                    continue

                coluna_real = None
                for c in json_df.columns:
                    if str(c).upper() == nome_coluna:
                        coluna_real = c
                        break

                if coluna_real is None:
                    continue

                serie_total = json_df[coluna_real].dropna().apply(normalizar_resposta)
                serie_total = serie_total[serie_total != ""]

                if serie_total.empty:
                    continue

                total_resp = len(serie_total)
                serie_valida = serie_total[~serie_total.apply(eh_sem_direcao)]
                sem_direcao = total_resp - len(serie_valida)

                pct_sem_direcao = round((sem_direcao / total_resp) * 100, 1) if total_resp else 0
                pct_definicao = round((len(serie_valida) / total_resp) * 100, 1) if total_resp else 0

                if serie_valida.empty:
                    continue

                contagem = serie_valida.value_counts().head(8).reset_index()
                contagem.columns = ["Resposta", "Entrevistas"]
                contagem["%"] = (contagem["Entrevistas"] / total_resp * 100).round(1)

                top1 = contagem.iloc[0]["Resposta"]
                top1_pct = contagem.iloc[0]["%"]

                perfis_respostas = []
                for _, r in contagem.head(3).iterrows():
                    resp = r["Resposta"]
                    fortes, fracos = grupos_fortes_fracos(df, nome_coluna, resp)

                    if fortes or fracos:
                        texto = f"{resp}: "
                        partes = []
                        if fortes:
                            partes.append("fortalezas - " + "; ".join(fortes))
                        if fracos:
                            partes.append("fragilidades - " + "; ".join(fracos))
                        texto += " | ".join(partes)
                        perfis_respostas.append(html.Li(texto))

                if tipo_pergunta == "ESTRATEGICA":
                    if len(contagem) >= 2:
                        segundo_resp = contagem.iloc[1]["Resposta"]
                        segundo_pct = contagem.iloc[1]["%"]
                        vantagem = round(top1_pct - segundo_pct, 1)

                        if vantagem <= 3:
                            status_disputa = "cenário tecnicamente muito competitivo"
                        elif vantagem <= 7:
                            status_disputa = "cenário competitivo, com liderança vulnerável"
                        elif vantagem <= 15:
                            status_disputa = "liderança moderada"
                        else:
                            status_disputa = "liderança ampla"

                        leitura = (
                            f"**Leitura estratégica:** {top1} aparece na liderança com {top1_pct}%, "
                            f"seguido por {segundo_resp}, com {segundo_pct}%. "
                            f"A vantagem atual é de {vantagem} pontos, indicando {status_disputa}. "
                            f"O índice de definição é de {pct_definicao}%, enquanto {pct_sem_direcao}% "
                            f"permanecem sem direção eleitoral/opinativa. "
                            f"Esse percentual de indefinição representa espaço real de crescimento, disputa e conversão eleitoral."
                        )
                    else:
                        leitura = (
                            f"**Leitura estratégica:** {top1} concentra {top1_pct}% das respostas. "
                            f"O índice de definição é de {pct_definicao}%."
                        )

                elif tipo_pergunta == "ANALITICA":
                    sentimentos = serie_total.apply(classificar_sentimento_resposta)
                    positivo = int((sentimentos == "POSITIVO").sum())
                    negativo = int((sentimentos == "NEGATIVO").sum())
                    neutro = int((sentimentos == "NEUTRO").sum())
                    positivo_pct = round((positivo / total_resp) * 100, 1) if total_resp else 0
                    negativo_pct = round((negativo / total_resp) * 100, 1) if total_resp else 0
                    neutro_pct = round((neutro / total_resp) * 100, 1) if total_resp else 0
                    saldo = round(positivo_pct - negativo_pct, 1)

                    if positivo or negativo or neutro:
                        leitura = (
                            f"**Leitura analítica:** As respostas positivas somam {positivo_pct}%, "
                            f"as negativas {negativo_pct}% e as neutras {neutro_pct}%. "
                            f"O saldo positivo/negativo é de {saldo} pontos. "
                            f"A resposta mais citada foi {top1}, com {top1_pct}%."
                        )
                    else:
                        leitura = (
                            f"**Leitura analítica:** A resposta mais citada foi {top1}, com {top1_pct}%. "
                            f"O índice de definição é de {pct_definicao}%, enquanto {pct_sem_direcao}% ficaram sem direção clara."
                        )

                else:
                    leitura = (
                        f"A resposta mais citada foi {top1}, com {top1_pct}%. "
                        f"O índice de definição é de {pct_definicao}%."
                    )

                tabela_top = tabela_padrao(contagem, 8)

                filhos_bloco = [
                    html.H4(
                        f"[{tipo_pergunta}] {label}",
                        style={"marginTop": "0", "color": "white"}
                    ),
                    html.Div([
                        card("Definição", f"{pct_definicao}%", "Respostas válidas/direcionais", "#0f766e"),
                        card("Sem direção", f"{pct_sem_direcao}%", "NS/NR, indecisos, branco/nulo", "#92400e"),
                        card("Top resposta", str(top1), f"{top1_pct}% do total", "#1d4ed8"),
                    ], style={
                        "display": "grid",
                        "gridTemplateColumns": "repeat(3, 1fr)",
                        "gap": "14px",
                        "marginBottom": "16px"
                    }),
                    tabela_top,
                    dcc.Markdown(leitura, style={
                        "fontSize": "15px",
                        "lineHeight": "1.6",
                        "color": "#cbd5e1",
                        "marginTop": "14px"
                    })
                ]

                if perfis_respostas:
                    filhos_bloco.append(
                        html.Div([
                            html.H4("Fortalezas e fragilidades", style={
                                "color": "white",
                                "marginTop": "16px"
                            }),
                            html.Ul(perfis_respostas, style={
                                "color": "#cbd5e1",
                                "lineHeight": "1.7"
                            })
                        ])
                    )

                blocos_perguntas.append(
                    html.Div(filhos_bloco, style={
                        "background": "#111827",
                        "borderRadius": "14px",
                        "border": "1px solid #1f2937",
                        "padding": "20px",
                        "marginBottom": "18px"
                    })
                )

            except Exception as e:
                blocos_perguntas.append(
                    html.Div(
                        f"Erro ao processar a pergunta '{pergunta.get('label', '')}': {e}",
                        style={
                            "background": "#7f1d1d",
                            "padding": "14px",
                            "borderRadius": "10px",
                            "color": "white",
                            "marginBottom": "12px"
                        }
                    )
                )

        resumo = (
            f"A pesquisa possui {total} entrevistas válidas. "
            f"As principais localidades da amostra são: {top_localidades_txt}. "
            f"A distribuição por sexo é: {sexo_txt}. "
            f"A distribuição por faixa etária é: {idade_txt}. "
            f"{mapa_txt}"
        )

        return html.Div([
            html.Div([
                card("Total de entrevistas", f"{total:,}".replace(",", "."), "Base válida"),
                card("Sexo predominante", sexo_top, f"{pct_fem}% feminino | {pct_masc}% masculino"),
                card("Faixa etária líder", idade_top, "Maior presença na amostra"),
                card("GPS válido", f"{gps_pct}%", f"{entrevistas_gps} entrevistas com mapa"),
            ], style={
                "display": "grid",
                "gridTemplateColumns": "repeat(4, 1fr)",
                "gap": "18px",
                "marginBottom": "22px"
            }),

            html.Div([
                html.H3("Resumo Executivo Automático", style={"marginTop": "0"}),
                html.P(resumo, style={
                    "fontSize": "16px",
                    "lineHeight": "1.6",
                    "color": "#e5e7eb"
                })
            ], style={
                "background": "#111827",
                "borderRadius": "14px",
                "border": "1px solid #1f2937",
                "padding": "20px",
                "marginBottom": "22px"
            }),

            html.Div([
                html.Div([
                    html.H3("Top Localidades da Amostra", style={"marginTop": "0"}),
                    tabela_localidades
                ], style={
                    "background": "#111827",
                    "borderRadius": "14px",
                    "border": "1px solid #1f2937",
                    "padding": "20px"
                }),
                html.Div([
                    html.H3("Distribuição por Sexo", style={"marginTop": "0"}),
                    tabela_sexo
                ], style={
                    "background": "#111827",
                    "borderRadius": "14px",
                    "border": "1px solid #1f2937",
                    "padding": "20px"
                }),
            ], style={
                "display": "grid",
                "gridTemplateColumns": "1fr 1fr",
                "gap": "18px",
                "marginBottom": "22px"
            }),

            html.Div([
                html.H3("Distribuição por Faixa Etária", style={"marginTop": "0"}),
                tabela_idade
            ], style={
                "background": "#111827",
                "borderRadius": "14px",
                "border": "1px solid #1f2937",
                "padding": "20px"
            }),

            html.Div([
                html.H3("Análise Inteligente das Perguntas", style={"marginTop": "0"}),
                html.Div(blocos_perguntas if blocos_perguntas else [
                    html.Div(
                        "Nenhuma pergunta válida encontrada para gerar inteligência.",
                        style={
                            "background": "#111827",
                            "padding": "20px",
                            "borderRadius": "14px",
                            "border": "1px solid #1f2937",
                            "color": "#cbd5e1"
                        }
                    )
                ])
            ], style={
                "background": "#0f172a",
                "borderRadius": "14px",
                "padding": "0px",
                "marginTop": "22px"
            })
        ])

    except Exception as e:
        return html.Div([
            html.H3("Erro ao gerar Inteligência", style={"color": "#fecaca"}),
            html.Pre(str(e), style={
                "background": "#7f1d1d",
                "color": "white",
                "padding": "16px",
                "borderRadius": "10px",
                "whiteSpace": "pre-wrap"
            })
        ])

ODK_URL = "https://app.ar7pesquisas.com.br"
ODK_USER = "augusto.estatistico@gmail.com"
ODK_PASS = "@Mat050dois"

@app.callback(
    Output("link-exportar-excel", "href"),
    Input("pesquisa", "value")
)
def atualizar_link_excel(pesquisa_id):

    if not pesquisa_id:
        return "#"

    return f"/exportar_excel/{pesquisa_id}"

@app.callback(
    Output("link-gerar-pdf", "href"),
    Input("pesquisa", "value")
)
def atualizar_link_pdf(pesquisa_id):

    if not pesquisa_id:
        return "#"

    return f"/gerar_pdf/{pesquisa_id}"


@app.callback(
    Output("link-exportar-tabelas", "href"),
    Input("pesquisa", "value")
)
def atualizar_link_tabelas(pesquisa_id):

    if not pesquisa_id:
        return "#"

    return f"/exportar_tabelas/{pesquisa_id}"


# =========================
# CALLBACK MENU POR SEÇÃO
# =========================
@app.callback(
    Output("secao-ativa", "data"),
    [
        Input("btn-visao-geral", "n_clicks"),
        Input("btn-entrevistas", "n_clicks"),
        Input("btn-mapas", "n_clicks"),
        Input("btn-perguntas", "n_clicks"),
        Input("btn-entrevistadores", "n_clicks"),
        Input("btn-relatorios", "n_clicks"),
        Input("btn-inteligencia", "n_clicks"),
        
    ]
)
def mudar_secao(n1, n2, n3, n4, n5, n6, n7):
    ctx = dash.callback_context

    if not ctx.triggered:
        return "visao-geral"

    botao = ctx.triggered[0]["prop_id"].split(".")[0]

    mapa = {
        "btn-visao-geral": "visao-geral",
        "btn-entrevistas": "entrevistas",
        "btn-mapas": "mapas",
        "btn-perguntas": "perguntas",
        "btn-entrevistadores": "entrevistadores",
        "btn-relatorios": "relatorios",
        "btn-inteligencia": "inteligencia",
    }

    return mapa.get(botao, "visao-geral")


@app.callback(
    [
        Output("secao-visao-geral", "style"),
        Output("secao-entrevistas", "style"),
        Output("secao-mapas", "style"),
        Output("secao-perguntas", "style"),
        Output("secao-entrevistadores", "style"),
        Output("secao-relatorios", "style"),
        Output("secao-inteligencia", "style"),
    ],
    Input("secao-ativa", "data")
)
def exibir_secao(secao):
    base = {"display": "block", "marginBottom": "20px"}
    oculto = {"display": "none"}

    style_entrevistas = {
        "display": "block",
        "background": "#111827",
        "borderRadius": "14px",
        "border": "1px solid #1f2937",
        "padding": "20px",
        "marginTop": "20px"
    }

    style_entrevistadores = {
        "display": "block",
        "background": "#111827",
        "borderRadius": "14px",
        "border": "1px solid #1f2937",
        "padding": "18px"
    }

    return (
        base if secao == "visao-geral" else oculto,
        style_entrevistas if secao == "entrevistas" else oculto,
        base if secao == "mapas" else oculto,
        base if secao == "perguntas" else oculto,
        style_entrevistadores if secao == "entrevistadores" else oculto,
        base if secao == "relatorios" else oculto,
        base if secao == "inteligencia" else oculto,
    )

# =========================
# AUDIO ENDPOINT
# =========================
@server.route("/audio/<int:pesquisa_id>/<path:submission_id>/<path:nome_arquivo>")
def ouvir_audio(pesquisa_id, submission_id, nome_arquivo):

    if "cliente_id" not in session:
        return "Não autorizado", 403

    df = pd.read_sql(
        text("""
            SELECT projeto_odk, form_id
            FROM pesquisas
            WHERE id = :pesquisa_id
        """),
        engine,
        params={"pesquisa_id": pesquisa_id}
    )

    if df.empty:
        return "Pesquisa não encontrada", 404

    projeto_odk = df.iloc[0]["projeto_odk"]
    form_id = df.iloc[0]["form_id"]

    import urllib.parse
    import tempfile
    import subprocess
    import requests
    from requests.auth import HTTPBasicAuth

    submission_encoded = urllib.parse.quote(submission_id, safe="")
    arquivo_encoded = urllib.parse.quote(nome_arquivo, safe="")

    url = (
        f"https://app.ar7pesquisas.com.br/v1/projects/{projeto_odk}"
        f"/forms/{form_id}"
        f"/submissions/{submission_encoded}"
        f"/attachments/{arquivo_encoded}"
    )

    r = requests.get(
        url,
        auth=HTTPBasicAuth(
            "augusto.estatistico@gmail.com",
            "@Mat050dois"
        )
    )

    if r.status_code != 200:
        return f"Erro ao buscar áudio: {r.status_code}", 500

    with tempfile.NamedTemporaryFile(suffix=".amr", delete=False) as entrada:
        entrada.write(r.content)
        entrada_path = entrada.name

    saida_path = entrada_path.replace(".amr", ".mp3")

    processo = subprocess.run(
        [
            "ffmpeg",
            "-y",
            "-i", entrada_path,
            "-acodec", "libmp3lame",
            "-ab", "64k",
            saida_path
        ],
        capture_output=True,
        text=True
    )

    if processo.returncode != 0:
        return f"Erro ao converter áudio: {processo.stderr}", 500

    with open(saida_path, "rb") as f:
        audio_mp3 = f.read()

    return Response(
        audio_mp3,
        mimetype="audio/mpeg"
    )
# =========================
# EXPORTAR EXCEL
# =========================
@server.route("/exportar_excel/<int:pesquisa_id>")
def exportar_excel(pesquisa_id):

    if "cliente_id" not in session:
        return "Não autorizado", 403

    df = carregar_dados(pesquisa_id)

    if df.empty:
        return "Sem dados", 404

    dados_json = df["dados"].apply(normalizar_json)
    json_df = pd.json_normalize(dados_json)

    df_export = pd.concat(
        [df.drop(columns=["dados"]), json_df],
        axis=1
    )

    import io

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="base")

    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=pesquisa_{pesquisa_id}.xlsx"
        }
    )

# =========================
# EXPORTAR TABELAS (% E AMOSTRA)
# =========================
@server.route("/exportar_tabelas/<int:pesquisa_id>")
def exportar_tabelas(pesquisa_id):

    if "cliente_id" not in session:
        return "Não autorizado", 403

    df = carregar_dados(pesquisa_id)

    if df.empty:
        return "Sem dados", 404

    try:
        pesquisa_nome = [
            opt["label"]
            for opt in lista_pesquisas(session.get("cliente_id"))
            if opt["value"] == pesquisa_id
        ][0]
    except Exception:
        pesquisa_nome = f"Pesquisa ID {pesquisa_id}"

    cliente = get_cliente(session.get("cliente_id"))

    dados_json = df["dados"].apply(normalizar_json)
    json_df = pd.json_normalize(dados_json)

    perguntas_df = pd.read_sql(
        text("""
            SELECT UPPER(name) AS name, label
            FROM perguntas_pesquisa
            WHERE pesquisa_id = :pesquisa_id
            AND exibir_dashboard = true
            ORDER BY id
        """),
        engine,
        params={"pesquisa_id": pesquisa_id}
    )

    def deve_excluir_tabela(nome, label):
        txt = f"{nome} {label}".upper()
        excluir = [
            "ENTREVISTADOR",
            "SUBMISSION",
            "_SYSTEM",
            "_ID",
            "INSTANCE",
            "UUID",
            "GPS",
            "LATITUDE",
            "LONGITUDE",
            "ACCURACY",
            "PESQUISA_INICIO",
            "PESQUISA_FIM",
            "HOJE",
            "NOME"
        ]
        return any(x in txt for x in excluir)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        resumo_rows = []
        tabelas_rows = []

        for _, pergunta in perguntas_df.iterrows():
            nome_coluna = str(pergunta["name"]).upper()
            label = str(pergunta["label"]).strip()

            if deve_excluir_tabela(nome_coluna, label):
                continue

            coluna_real = None

            for c in json_df.columns:
                if str(c).upper() == nome_coluna:
                    coluna_real = c
                    break

            if coluna_real is None:
                continue

            serie = json_df[coluna_real].dropna().apply(normalizar_resposta)
            serie = serie[serie != ""]

            if serie.empty:
                continue

            contagem = serie.value_counts(dropna=False).reset_index()
            contagem.columns = ["Resposta", "Amostra"]
            base = int(contagem["Amostra"].sum())

            if base == 0:
                continue

            contagem["%"] = (contagem["Amostra"] / base).round(4)
            contagem = contagem[["Resposta", "%", "Amostra"]]

            top_resp = str(contagem.iloc[0]["Resposta"])
            top_pct = float(contagem.iloc[0]["%"])

            resumo_rows.append({
                "Pergunta": label,
                "Base": base,
                "Top resposta": top_resp,
                "% Top": top_pct
            })

            tabelas_rows.append([label, "", ""])
            tabelas_rows.append(["Row Labels", "%", "Amostra"])

            for _, row in contagem.iterrows():
                tabelas_rows.append([
                    row["Resposta"],
                    float(row["%"]),
                    int(row["Amostra"])
                ])

            tabelas_rows.append(["Total Geral", 1.0, base])
            tabelas_rows.append(["", "", ""])
            tabelas_rows.append(["", "", ""])

        if resumo_rows:
            pd.DataFrame(resumo_rows).to_excel(
                writer,
                index=False,
                sheet_name="Resumo"
            )
        else:
            pd.DataFrame([{
                "Pergunta": "Nenhuma pergunta encontrada",
                "Base": 0,
                "Top resposta": "",
                "% Top": 0
            }]).to_excel(
                writer,
                index=False,
                sheet_name="Resumo"
            )

        if tabelas_rows:
            pd.DataFrame(tabelas_rows).to_excel(
                writer,
                index=False,
                header=False,
                sheet_name="Tabelas"
            )
        else:
            pd.DataFrame([["Nenhuma tabela encontrada", "", ""]]).to_excel(
                writer,
                index=False,
                header=False,
                sheet_name="Tabelas"
            )

        wb_excel = writer.book

        for ws in wb_excel.worksheets:
            ws.freeze_panes = "A2"

        ws_resumo = wb_excel["Resumo"]
        ws_tabelas = wb_excel["Tabelas"]

        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        azul = "1F4E78"
        cinza = "D9EAF7"
        borda = Side(style="thin", color="B7B7B7")

        # Estilo resumo
        for cell in ws_resumo[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=azul)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(top=borda, left=borda, right=borda, bottom=borda)

        for row in ws_resumo.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(top=borda, left=borda, right=borda, bottom=borda)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for cell in ws_resumo["D"][1:]:
            cell.number_format = "0.0%"

        ws_resumo.column_dimensions["A"].width = 70
        ws_resumo.column_dimensions["B"].width = 12
        ws_resumo.column_dimensions["C"].width = 28
        ws_resumo.column_dimensions["D"].width = 12

        # Estilo tabelas
        for row in ws_tabelas.iter_rows():
            first = row[0].value if len(row) else None

            if first and first not in ["Row Labels", "Total Geral"]:
                # Linha de título da pergunta
                row[0].font = Font(bold=True, color="FFFFFF", size=11)
                row[0].fill = PatternFill("solid", fgColor=azul)
                row[0].alignment = Alignment(wrap_text=True)
                for cell in row:
                    cell.border = Border(top=borda, left=borda, right=borda, bottom=borda)

            elif first == "Row Labels":
                for cell in row:
                    cell.font = Font(bold=True, color="000000")
                    cell.fill = PatternFill("solid", fgColor=cinza)
                    cell.border = Border(top=borda, left=borda, right=borda, bottom=borda)
                    cell.alignment = Alignment(horizontal="center")

            elif first == "Total Geral":
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="E2F0D9")
                    cell.border = Border(top=borda, left=borda, right=borda, bottom=borda)

            else:
                for cell in row:
                    if cell.value not in [None, ""]:
                        cell.border = Border(top=borda, left=borda, right=borda, bottom=borda)

        for cell in ws_tabelas["B"]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0%"

        ws_tabelas.column_dimensions["A"].width = 65
        ws_tabelas.column_dimensions["B"].width = 12
        ws_tabelas.column_dimensions["C"].width = 12

        # Aba de identificação
        ws_info = wb_excel.create_sheet("Info", 0)
        ws_info["A1"] = "Cliente"
        ws_info["B1"] = cliente.get("nome", "")
        ws_info["A2"] = "Pesquisa"
        ws_info["B2"] = pesquisa_nome
        ws_info["A3"] = "Total de entrevistas na base"
        ws_info["B3"] = len(df)
        ws_info["A4"] = "Observação"
        ws_info["B4"] = "As tabelas consideram todas as respostas informadas, incluindo Não sabe, indecisos, branco/nulo, NS/NR, nenhum e demais respostas indefinidas."
        for cell in ws_info["A"]:
            cell.font = Font(bold=True)
        ws_info.column_dimensions["A"].width = 28
        ws_info.column_dimensions["B"].width = 90
        ws_info["B4"].alignment = Alignment(wrap_text=True)

    output.seek(0)

    nome_arquivo = (
        f"tabelas_percentual_amostra_{pesquisa_id}.xlsx"
    )

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={nome_arquivo}"
        }
    )


# =========================
# GERAR PDF
# =========================
@server.route("/gerar_pdf/<int:pesquisa_id>")
def gerar_pdf(pesquisa_id):

    if "cliente_id" not in session:
        return "Não autorizado", 403

    df = carregar_dados(pesquisa_id)

    if df.empty:
        return "Sem dados", 404

    cliente = get_cliente(session.get("cliente_id"))

    try:
        pesquisa_nome = [
            opt["label"]
            for opt in lista_pesquisas(session.get("cliente_id"))
            if opt["value"] == pesquisa_id
        ][0]
    except:
        pesquisa_nome = f"Pesquisa ID {pesquisa_id}"

    dados_json = df["dados"].apply(normalizar_json)
    json_df = pd.json_normalize(dados_json)

    perguntas_df = pd.read_sql(
        text("""
            SELECT UPPER(name) AS name, label
            FROM perguntas_pesquisa
            WHERE pesquisa_id = :pesquisa_id
            AND exibir_dashboard = true
            ORDER BY id
        """),
        engine,
        params={"pesquisa_id": pesquisa_id}
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    styles = getSampleStyleSheet()
    elementos = []
    elementos.append(Paragraph("<b>RELATÓRIO DE PESQUISA</b>", styles["Title"]))
    elementos.append(Spacer(1, 18))
    elementos.append(Paragraph(f"<b>Cliente:</b> {cliente['nome']}", styles["Normal"]))
    elementos.append(Paragraph(f"<b>Pesquisa:</b> {pesquisa_nome}", styles["Normal"]))
    elementos.append(Paragraph(f"<b>Total de entrevistas:</b> {len(df)}", styles["Normal"]))
    elementos.append(Spacer(1, 24))

    elementos.append(Paragraph("<b>1. Resumo Executivo</b>", styles["Heading1"]))

    total = len(df)
    masc = int((df["sexo"] == "Masculino").sum())
    fem = int((df["sexo"] == "Feminino").sum())
    entrevistadores = df["entrevistador"].nunique()

    masc_pct = round(masc / total * 100, 1) if total else 0
    fem_pct = round(fem / total * 100, 1) if total else 0

    elementos.append(Paragraph(f"Entrevistas realizadas: <b>{total}</b>", styles["Normal"]))
    elementos.append(Paragraph(f"Masculino: <b>{masc}</b> entrevistas ({masc_pct}%)", styles["Normal"]))
    elementos.append(Paragraph(f"Feminino: <b>{fem}</b> entrevistas ({fem_pct}%)", styles["Normal"]))
    elementos.append(Paragraph(f"Entrevistadores em campo: <b>{entrevistadores}</b>", styles["Normal"]))
    elementos.append(Spacer(1, 18))

    elementos.append(Paragraph("<b>2. Perfil da Amostra</b>", styles["Heading1"]))

    for coluna, titulo in [
        ("sexo", "Sexo"),
        ("idade", "Faixa Etária"),
        ("localidade", "Localidade")
    ]:
        elementos.append(Paragraph(f"<b>{titulo}</b>", styles["Heading2"]))

        resumo = df[coluna].value_counts().reset_index()
        resumo.columns = ["Categoria", "Quantidade"]

        for _, row in resumo.iterrows():
            pct = round(row["Quantidade"] / total * 100, 1)
            elementos.append(
                Paragraph(
                    f"{row['Categoria']}: {row['Quantidade']} ({pct}%)",
                    styles["Normal"]
                )
            )

        elementos.append(Spacer(1, 10))

    elementos.append(PageBreak())

    elementos.append(Paragraph("<b>3. Resultados das Perguntas</b>", styles["Heading1"]))

    for _, pergunta in perguntas_df.iterrows():

        nome_coluna = pergunta["name"]
        label = pergunta["label"]

        coluna_real = None

        for c in json_df.columns:
            if c.upper() == nome_coluna:
                coluna_real = c
                break

        if coluna_real is None:
            continue

        serie = json_df[coluna_real].dropna().astype(str).str.strip()
        serie = serie[serie != ""]

        if serie.empty:
            continue

        elementos.append(Paragraph(f"<b>{label}</b>", styles["Heading2"]))

        serie_valida = serie[~serie.apply(eh_sem_direcao)]

        if serie_valida.empty:
            continue

        contagem = serie_valida.value_counts().reset_index()

                
        contagem.columns = ["Resposta", "Quantidade"]
        base = contagem["Quantidade"].sum()

        serie_valida = serie[~serie.apply(eh_sem_direcao)]
        sem_direcao = len(serie) - len(serie_valida)

        pct_definicao = round((len(serie_valida) / len(serie)) * 100, 1) if len(serie) else 0
        pct_sem_direcao = round((sem_direcao / len(serie)) * 100, 1) if len(serie) else 0

        elementos.append(
            Paragraph(
                f"<b>Índice de definição:</b> {pct_definicao}% | "
                f"<b>Sem direção:</b> {pct_sem_direcao}%",
                styles["Normal"]
            )
        )        

        for _, row in contagem.iterrows():
            pct = round(row["Quantidade"] / base * 100, 1)
            elementos.append(
                Paragraph(
                    f"{row['Resposta']}: {row['Quantidade']} ({pct}%)",
                    styles["Normal"]
                )
            )



        perfis_respostas = []

        for _, r in contagem.head(3).iterrows():
            resp = r["Resposta"]
            perfil = perfil_resposta(df, nome_coluna, resp)

            if perfil:
                perfis_respostas.append(
                    html.Li(
                        f"{resp}: perfil predominante {perfil['sexo']}, "
                        f"{perfil['idade']}, maior presença em {perfil['localidade']}."
                    )
                )

        if not contagem.empty:

            top_resp = contagem.iloc[0]["Resposta"]

            top_qtd = contagem.iloc[0]["Quantidade"]

            top_pct = round(
                (top_qtd / base) * 100,
                1
            ) if base else 0

            destaques = grupos_fortes(df, nome_coluna, top_resp)
            destaque_txt = "; ".join(destaques) if destaques else "sem destaque relevante acima da média"


            elementos.append(
                Paragraph(
                    f"<b>Leitura automática:</b> "
                    f"A resposta mais citada foi <b>{top_resp}</b>, "
                    f"com {top_pct}% das respostas. "
                    f"{'Grupos e territórios com destaque: ' + destaque_txt + '.' if destaques else ''}",
                    styles["Normal"]
                )
            )

            elementos.append(Spacer(1, 8))                
                
        elementos.append(Spacer(1, 14))

    elementos.append(PageBreak())

    elementos.append(Paragraph("<b>4. Observação Técnica</b>", styles["Heading1"]))
    elementos.append(Paragraph(
        "Relatório gerado automaticamente pela plataforma Ipsensus Survey, "
        "com base nos dados coletados via ODK, processados no banco Supabase "
        "e visualizados no dashboard operacional.",
        styles["Normal"]
    ))

    doc.build(elementos)

    buffer.seek(0)

    return Response(
        buffer.getvalue(),
        mimetype="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=relatorio_{pesquisa_id}.pdf"
        }
    )


# =========================
# ETL ENDPOINT
# =========================
@server.route("/etl", methods=["GET", "HEAD"])
def etl():

    token = request.args.get("token")

    if token != os.getenv("ETL_TOKEN", "123456"):
        return "Token inválido", 403

    # Executa tanto para GET quanto para HEAD
    subprocess.Popen(
        [sys.executable, "etl.py"],
        cwd=os.path.dirname(os.path.abspath(__file__))
    )

    if request.method == "HEAD":
        return "", 200

    return "ETL iniciado em segundo plano", 200

# =========================
# RUN
# =========================
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8050))
    app.run(host="0.0.0.0", port=port)
